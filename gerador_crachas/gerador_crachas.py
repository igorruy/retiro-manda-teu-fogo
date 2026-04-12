#!/usr/bin/env python3
"""
Gerador de Crachás - Retiro de Crisma 2026 "Manda Teu Fogo"
Lê lista de pessoas de um Excel e gera PDF A4 com crachás prontos para corte.

Uso:
    python gerar_crachas.py                            # usa lista_crachas.xlsx padrão
    python gerar_crachas.py minha_lista.xlsx           # usa outro arquivo Excel
    python gerar_crachas.py lista.xlsx saida.pdf       # define nome da saída

Colunas obrigatórias no Excel:
    Nome  |  Equipe

A pasta 'design/' deve estar no mesmo diretório com os arquivos:
    guardioes.png, missao.png, providencia.png, movimento.png,
    intersessao.png, saude.png, secretaria.png, servos_do_altar.png,
    fortaleza.png, entendimento.png, ciencia.png, conselho.png,
    sabedoria.png, piedade.png, temor_a_deus.png
"""

import sys
import os
from pathlib import Path
from collections import defaultdict

import pandas as pd
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor, black, white, Color


# ─── Configurações ────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).parent
DESIGN_DIR = SCRIPT_DIR / "design"

# Dimensões do crachá (cm → mm)
CRACHA_W = 95   # mm
CRACHA_H = 65   # mm

# Margem da página
PAGE_MARGIN = 5  # mm

# Área de texto do nome dentro do crachá (em fração 0–1 da altura)
# Entre o fim do logo e o início da faixa cinza
NAME_AREA_TOP    = 0.30   # 30% do topo
NAME_AREA_BOTTOM = 0.848  # 84.8% do topo (início da faixa cinza)

# Estilo do texto do nome
NAME_FONT      = "Helvetica-Bold"
NAME_COLOR     = HexColor("#1a1a1a")
MAX_FONT_PT    = 20
MIN_FONT_PT    = 7

# Borda de corte (desenhada em cada crachá)
CUT_BORDER_COLOR = Color(0.5, 0.5, 0.5, alpha=1.0)
CUT_BORDER_WIDTH = 0.5   # pts — ≈ 0.18 mm, visível mas discreto

# Mapeamento equipe → arquivo de design
EQUIPE_MAP = {
    "guardiões":      "gradioes.png",
    "guardioes":      "gradioes.png",
    "missão":         "missao.png",
    "missao":         "missao.png",
    "providência":    "providencia.png",
    "providencia":    "providencia.png",
    "movimento":      "movimento.png",
    "intercessão":    "intersessao.png",
    "intersessão":    "intersessao.png",
    "intercessao":    "intersessao.png",
    "intersessao":    "intersessao.png",
    "saúde":          "saude.png",
    "saude":          "saude.png",
    "secretaria":     "secretaria.png",
    "servos do altar":"servos_do_altar.png",
    "servos_do_altar":"servos_do_altar.png",
    "fortaleza":      "fortaleza.png",
    "entendimento":   "entendimento.png",
    "ciência":        "ciencia.png",
    "ciencia":        "ciencia.png",
    "conselho":       "conselho.png",
    "sabedoria":      "sabedoria.png",
    "piedade":        "piedade.png",
    "temor a deus":   "temor_a_deus.png",
    "temor_a_deus":   "temor_a_deus.png",
}


# ─── Funções auxiliares ───────────────────────────────────────────────────────

def load_data(excel_path: str) -> pd.DataFrame:
    df = pd.read_excel(excel_path)
    df.columns = [c.strip() for c in df.columns]

    col_map = {}
    for col in df.columns:
        low = col.lower()
        if "nome" in low:     col_map[col] = "Nome"
        elif "equipe" in low: col_map[col] = "Equipe"
    df = df.rename(columns=col_map)

    missing = {"Nome", "Equipe"} - set(df.columns)
    if missing:
        raise ValueError(f"Colunas não encontradas: {missing}\nColunas disponíveis: {list(df.columns)}")

    df = df.dropna(subset=["Nome", "Equipe"])
    df["Nome"]   = df["Nome"].astype(str).str.strip()
    df["Equipe"] = df["Equipe"].astype(str).str.strip()
    return df


def resolve_template(equipe: str) -> Path | None:
    key = equipe.lower().strip()
    filename = EQUIPE_MAP.get(key)
    if filename:
        p = DESIGN_DIR / filename
        if p.exists():
            return p
    # Busca fuzzy: procura arquivo cujo nome contenha parte do nome da equipe
    for f in DESIGN_DIR.glob("*.png"):
        if key.replace(" ", "_") in f.stem or f.stem in key.replace(" ", "_"):
            return f
    return None


def wrap_name(c: canvas.Canvas, name: str, pt: int, max_w: float) -> list:
    """Quebra o nome em linhas que caibam na largura, mantendo fonte fixa."""
    c.setFont(NAME_FONT, pt)
    words = name.split()
    lines = []
    current = ""
    for word in words:
        test = (current + " " + word).strip()
        if c.stringWidth(test, NAME_FONT, pt) <= max_w:
            current = test
        else:
            if current:
                lines.append(current)
            current = word
    if current:
        lines.append(current)
    return lines


def draw_cracha(c: canvas.Canvas, template: Path | None, name: str,
                x: float, y: float, w: float, h: float):
    """
    Desenha um crachá na posição (x, y) — canto inferior esquerdo em pontos ReportLab.
    w, h em pontos.
    """
    # Template de fundo
    if template and template.exists():
        c.drawImage(str(template), x, y, width=w, height=h,
                    preserveAspectRatio=False, mask="auto")
    else:
        # Fallback: retângulo branco com bordas
        c.setFillColor(white)
        c.rect(x, y, w, h, fill=1, stroke=0)

    # Área do nome em coordenadas ReportLab (Y cresce para cima)
    area_top_y    = y + h * (1 - NAME_AREA_TOP)
    area_bottom_y = y + h * (1 - NAME_AREA_BOTTOM)
    area_h = area_top_y - area_bottom_y
    area_w = w * 0.85
    center_x = x + w / 2
    center_y = (area_top_y + area_bottom_y) / 2

    pt = MAX_FONT_PT
    line_h = pt * 1.25
    lines = wrap_name(c, name, pt, area_w)
    # Se as linhas não cabem na altura, reduz fonte como fallback
    while len(lines) * line_h > area_h and pt > MIN_FONT_PT:
        pt -= 1
        line_h = pt * 1.25
        lines = wrap_name(c, name, pt, area_w)

    c.setFont(NAME_FONT, pt)
    c.setFillColor(NAME_COLOR)
    total_h = len(lines) * line_h
    start_y = center_y + total_h / 2 - line_h * 0.75
    for line in lines:
        c.drawCentredString(center_x, start_y, line)
        start_y -= line_h

    # Borda de corte fina (0.5pt ≈ 0.18mm) desenhada por cima da imagem
    c.saveState()
    c.setStrokeColor(CUT_BORDER_COLOR)
    c.setLineWidth(CUT_BORDER_WIDTH)
    c.rect(x, y, w, h, fill=0, stroke=1)
    c.restoreState()




def generate_pdf(df: pd.DataFrame, output_path: str):
    A4_W, A4_H = A4
    margin = PAGE_MARGIN * mm
    cw = CRACHA_W * mm
    ch = CRACHA_H * mm

    cols = int((A4_W - 2 * margin) / cw)
    rows = int((A4_H - 2 * margin) / ch)
    per_page = cols * rows

    # Origem (canto inferior esquerdo da grade de crachás)
    x0 = (A4_W - cols * cw) / 2
    y0 = (A4_H - rows * ch) / 2

    records = list(df.itertuples(index=False))
    total = len(records)
    missing_templates = set()

    c = canvas.Canvas(output_path, pagesize=A4)

    for page_start in range(0, total, per_page):
        batch = records[page_start:page_start + per_page]

        for idx, person in enumerate(batch):
            col = idx % cols
            row = idx // cols
            # ReportLab Y=0 embaixo; posicionar da linha superior
            badge_x = x0 + col * cw
            badge_y = y0 + (rows - 1 - row) * ch

            template = resolve_template(person.Equipe)
            if template is None:
                missing_templates.add(person.Equipe)

            draw_cracha(c, template, person.Nome, badge_x, badge_y, cw, ch)

        c.showPage()

    c.save()

    pages = -(-total // per_page)
    print(f"✅  PDF gerado: {output_path}")
    print(f"    {total} crachás | {cols}×{rows} por página | {pages} página(s)")
    if missing_templates:
        print(f"⚠️  Templates não encontrados para: {missing_templates}")
        print(f"    Verifique a pasta 'design/' e os nomes das equipes.")


# ─── Excel modelo ─────────────────────────────────────────────────────────────

def create_sample_excel(path: str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Crachás"

    headers = ["Nome", "Equipe"]
    fill = PatternFill("solid", start_color="FF8C00")
    font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    exemplos = [
        ("Ana Paula Souza",   "Guardiões"),
        ("Carlos Eduardo",    "Guardiões"),
        ("Maria Fernanda",    "Missão"),
        ("João Victor",       "Providência"),
        ("Isabela Carvalho",  "Movimento"),
        ("Pedro Henrique",    "Secretaria"),
        ("Luiza Santos",      "Intercessão"),
        ("Rafael Oliveira",   "Saúde"),
        ("Beatriz Lima",      "Servos do Altar"),
        ("Gabriel Martins",   "Fortaleza"),
        ("Sophia Pereira",    "Entendimento"),
        ("Matheus Costa",     "Ciência"),
        ("Laura Alves",       "Conselho"),
        ("Enzo Ribeiro",      "Sabedoria"),
        ("Valentina Gomes",   "Piedade"),
        ("Arthur Ferreira",   "Temor a Deus"),
    ]
    for row, (nome, equipe) in enumerate(exemplos, 2):
        ws.cell(row=row, column=1, value=nome)
        ws.cell(row=row, column=2, value=equipe)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    wb.save(path)
    print(f"📊  Planilha modelo criada: {path}")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    excel_path  = sys.argv[1] if len(sys.argv) > 1 else "lista_crachas.xlsx"
    output_path = sys.argv[2] if len(sys.argv) > 2 else "crachas.pdf"

    # Cria planilha modelo se não existir
    if not os.path.exists(excel_path):
        print(f"📋  '{excel_path}' não encontrado — criando planilha modelo...")
        create_sample_excel(excel_path)

    if not DESIGN_DIR.exists():
        print(f"⚠️  Pasta 'design/' não encontrada em {SCRIPT_DIR}")
        print("    Crie a pasta e adicione os PNGs das equipes.")

    print(f"📋  Lendo: {excel_path}")
    df = load_data(excel_path)
    print(f"    {len(df)} pessoas encontradas")

    generate_pdf(df, output_path)


if __name__ == "__main__":
    main()